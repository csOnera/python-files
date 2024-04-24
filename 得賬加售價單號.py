import pathlib
import openpyxl

wb_path = input('交易成功 file path\ne.g. xxx.xlsx: ')

addingExcel = openpyxl.load_workbook(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\得物對賬/" + wb_path)
ws = addingExcel.active

folderP = pathlib.Path(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\得物對賬/" + input('對賬單 folder name: '))
folder = list(folderP.iterdir())

def searching(ref):
    for row in range(2, ws.max_row):
        if ws["c" + str(row)].value == ref:
            return row
    return False


for file in folder:
    filewb = openpyxl.load_workbook(file)
    filews = filewb['鉴定通过订单']

    # for checking any not in 交易成功
    fileNumberOfItem = filews.max_row - 3 #possibly 倉儲費
    notFoundCount = 0
    notFoundList = []

    for row in range(4,filews.max_row + 1):
        searchResult = searching(filews["b" + str(row)].value)
        if searchResult:
            ws["w" + str(searchResult)].value = file.name
            # ADD hkd price automatically
            ws["l" + str(searchResult)].value = filews['al' + str(row)].value
        elif filews["b" + str(row)].value != None:
            # print(filews["b" + str(row)].value + " not found")
            notFoundCount += 1
            notFoundList.append(filews["b" + str(row)].value)
        else:
            print("none type detected")
    print(file.name, "notFoundNumber訂單號找不到數量 (包括倉儲費):", notFoundCount, "下為找不到的訂單號")
    for notFound in notFoundList:
        print(notFound)

    filewb.save(file)


addingExcel.save(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\得物對賬/" + wb_path)
