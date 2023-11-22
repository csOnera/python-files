# import tesserocr

# from PIL import Image

# image = Image.open(r"D:\Downloads\code.jpg")

# image = image.convert('L')
# #这个是二值化阈值
# threshold = 150   
# table = []

# for i in  range(256):
#     if i < threshold:
#         table.append(0)
#     else:
#         table.append(1)
# #通过表格转换成二进制图片，1的作用是白色，不然就全部黑色了
# image = image.point(table,"1")
# # image.show()



# result = tesserocr.image_to_text(image)
# print(result)

# from datetime import datetime

# print(datetime.today().strftime('%d/%m/%Y'))

# dictlist = [{'a':1, 'b':2}]
# print(dictlist[0].values())

import datetime

# print(datetime.datetime.strptime("21/12/2008", "%d/%m/%Y").strftime("%Y-%m-%d"))


# import win32com.client
# xl = win32com.client.Dispatch("Excel.Application")
# xl.Visible = True

# pop = xl.Workbooks.Open(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\出貨OR退貨紀錄\POP出貨記錄LATEST VERSION-DESKTOP-833R29B.xlsx")


import mysql.connector
import openpyxl
import win32com.client
import os


# connection = mysql.connector.connect(
#     host='localhost',
#     port='3306',
#     user='root',
#     password='jdysz',
#     database='trial_database'
# )

# cursor = connection.cursor()

# cursor.execute("select * from refInvoiceNo where 型號 like 'T109%';")

# print(cursor.fetchall())

de = True

# print('hello {}'.format(0 if de == True else 1))
import re

# stri = "雪铁纳瑞士表 DS PH200M系列 经典复刻男表200M防水C036.407.36.050.00"

# # print(re.search('[A-Z]+\d+[.][\d.]+' ,stri).group())

# import openpyxl
# import re

# file = openpyxl.load_workbook(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\2020京東自營.xlsx")
# ws = file.active

# for row in range(1,ws.max_row + 1):
#     if ws["j" + str(row)].value != None:
#         if len(str(ws["j" + str(row)].value)) >20:
#             if re.search('[A-Z]+\d+[.][\d.]+' , str(ws["j" + str(row)].value)) != None and ws["k" + str(row)].value == None:
#                 print(re.search('[A-Z]+\d+[.][\d.]+' , str(ws["j" + str(row)].value)).group())
#                 ws["k" + str(row)].value = re.search('[A-Z]+\d+[.][\d.]+' , str(ws["j" + str(row)].value)).group()
# file.save(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\2020京東自營.xlsx")

# from selenium import webdriver
# from selenium.webdriver.chrome.service import Service
# from webdriver_manager.chrome import ChromeDriverManager

# driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
# service_obj = Service(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\DATABASE_TRIAL\python files\chromedriver.exe")
# driver = webdriver.Chrome(service=service_obj)
# driver.get("https://passport.shop.jd.com/")

# while True:
#     pass

# import pathlib


# folder = pathlib.Path(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\銷貨成本\2022-2023 銷貨成本\onera to cs invoice")
# flist = list(folder.iterdir())

# count = 0
# suma = 0

# for file in flist:
#     excel = openpyxl.load_workbook(file, data_only=True)
#     ws = excel.active

#     for row in range(1, ws.max_row):
#         if ws["g" + str(row)].value == "HKD":
#             count += 1
#             suma += float(ws["h" + str(row)].value)
#             print(ws["h9"].value, float(ws["h" + str(row)].value))

# print(count, suma)

# import openpyxl
# import re

# excel = openpyxl.load_workbook(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\以前紀錄\以往紀錄全集合 (2022年4月庫存至2023年3月尾庫存).xlsx")
# ws = excel['紀錄計數量']


# for row in range(1915, 2007):
    
#     net = "="

#     # for col in range(7, 8 + 19):
#     #     if type(ws.cell(row=row,column=col).value) == str:
#     #         # print(ws.cell(row=row,column=col).value)
#     #         if re.search("[+]\d+|[-]\d+",ws.cell(row=row,column=col).value) != None:
#     #             x = re.search("[+]\d+|[-]\d+",ws.cell(row=row,column=col).value).group()
#     #             net += x
#     #             ws["b" + str(row)].value = net
#     #             print(net)
#     f = ws.cell(row=row, column=6).value
#     if type(f) == str:
#         if re.search("=\d+",f) != None:
#             ws["c" + str(row)].value = re.search("=\d+",f).group()[1:]


# excel.save(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\以前紀錄\以往紀錄全集合 (2022年4月庫存至2023年3月尾庫存).xlsx")

# import re

# l = ["卡西欧（CASIO）G-SHOCK运动防震防水蓝牙运动时尚男表 GBA-800-1A", "斯沃琪（Swatch）瑞士手表男女腕表 原创系列 运动潮流休闲学生腕表 GB274", "斯沃琪（Swatch）瑞士手表男女腕表 原创系列 运动潮流休闲学生腕表 GW411", "天梭（TISSOT）瑞士男士手表 魅时系列 经典休闲轻奢腕表 T109.610.36.031.00"]
# rel = []
# for item in l:
#     rel.append(re.search("[a-zA-Z.\d\-]+\Z", item).group())
    
# print(rel)

# realOneList = [2200622, 2200626, 2200633, 'XK-20220318-0132', 'P24840', 'P24879', 'VLS182/2022', 'D07152', 'D07153', 'D07160', 'S262', 'S263', 'S264', 'D07157', '548df491-9404-09e4-d226-2899e312810d', 'VL167/2022', 'VL235/2022', 'D-00005(000210753)', 'D-00005(000210760)', 2200658, 2200662, 2200672, 2200673, 2200682, 2200684, 'VL144/2022', '黃埔交收單', 2200708, '33565563-cf35-2ff2-e901-e434ea3b1abd', 'D07274', 'S266', 'S267', 'D07273', 'D07290', 'VL181/2022', 'VL265/2022', 'VL266/2022', 'D07297', 'D07306', 'VLS258/2022', 'VLS266/2022', 'VL262/2022', 'VLS274/2022', 'VLS276/2022', 'D07328', 'D07329', 'VL302/2022', 'S268', 'b631e20a-cc59-c85f-bb1e-1753d03cb2b2', 'D07333', 'D07341', 'D07344', 'D07346', 'D07348', 'D07350', 'D07355', 'D07356', 'VLS297', 'S269', 'c50f5815-e14b-59c8-213a-b6bc4adfd469', 'D07385', 'VL358/2022', 'D07409', 'D07419', 'D07427', 'D07434', 'P25130', 'VLS393/2022', 'S270', 'S271', 'P25271', 'VL389/2022', 'S272', 'D07492', 'P25132-R01', 'VL388/2022', 'VL390/2022', 'VLS432/2022', 'VLS434/2022', '913a1f64-d818-db00-5a25-54b37efbdd46', 'VLS467/2022', 'P25274', 'P25133-R01', 'P25281-R01', 'VL456/2022', 'P25206-R01', 'D07598', 'S274', 'SI2211028', 'S275 CHECKed', 20221103, 'P25448', 'D07670', 'DRAGON', 'P25447', 'P25526', 'SI2212024', 'SI2212038', 'P25518', 'P25504-R01', 'P25559', 'P25519', 'SI2302013', 'S281', 'P25623', 'P25624', '8aef2e01-f1a8-9bf5-3e00-b2bf37737fe9', '290d05f0-80f2-ebe6-c975-65e30c61b9a0', 'LEE NGAI', 'SO2303029', 'P25646', 'SI2303023', 'P25538', 'D-00005 (000211278)', '148175e7-6178-b6ab-4a7d-d3d071f0cf6b']
# realCsList = ['D07133', 'D07129', 225322, 'D07136', 225323, 2200606, 'D07135', 2200607, 2200602, 'D07144', 'D07145', 'VLS156', 225329, 'D07148', 'U2203003', 'U2203034', 'D07138', 'D07161', 'SO2204004', 'D07166', 'D07170', 2200642, 225346, 'D07182', 'D07178', 'D07183', 225349, 'afa0fa93-57b3-7a4d-9a41-eae0e10a2480', 'VLS186/2022', 2200668, 225353, 'D07195', 'D07198', 'VLS200/2022', 'VLS202/2022', 'D07208', 225356, 'P24884', 2200688, 'D07179', 'D07212', '20220510-1 (CHEUNGMING)', 'D07235', 'D07202', 'D07239', 'VLS208/2022', 'P24886-R01', 'VLS223/2022', 'D07250', 'S265', 'D07244', 'P24798', 'D07257', 'D07267', 'SO2206002', 'D07280', 'VLS246/2022', 'VLS248/2022', 'P24990', 'D07309', 'VLS265/2022', 'SO2206013', 'VLS277/2022', 'VLS278/2022', 'D07324', 'SI2207003', 'VLS289/2022', 'SO2207009', 'VLS301/2022', 'VLS302/2022', 'VLS305/2022', 'PI7484', 'D07359', 'D07365', 'VLS318/2022', 'VLS320/2022', 'D07367', 'D07369', 'P25128-R01', 'D07374', 'D07376', 'SO2207018', 'D07391', 'VLS330/2022', 'VLS341/2022', 'D07396', 'P25084-R01', 'D07399', 'D07401', 'D07404', 'VLS343/2022', 'VLS355/2022', 'VLS346/2022', 'VLS344/2022', 'VLS350/2022', 'VLS359/2022', 'VL365/2022', 'VLS374/2022', 'D07446', 'D07456', 'VLS387/2022', 'VLS388/2022', 'VLS391/2022', 'VL387/2022', 'D07487', 'D07486', 'D07494', 'D07513', 'D07516', 'D07523', 'D07535', 'VLS458/2022', 'D07538', 'S273', 'VL432/2022', 'D07562', 'D07559', 'P25312', 'D07579', '096/2022', 'D07601', 'D07602', 'D07646', 'D07645', '106/2022', 'SO2211037 CHECKED', 'P25377-R02 CHECKED', 'P25382 CHECKED', 'D07668', 'D07669 CORRECTED', 'VL453/2022', 'VL454/2022', 'VL495/2022', 'VL485/2022', 'VL498/2022', 'VL499/2022', 'SI2211044', 'D07683', 'D07684', 'D07693', 'S277 (PHYSICALLY IN STOCK)', 'S278', 'SO2212020', 'SO2212031', 'cd9f0716-876a-6bda-9068-adb31eece8ef (已入庫 PHYSICAL AND RECORD', 'D07707', 'H2022122901', '002/2023', 'SI2301003', 'D07747', 'S280', 'S279', 'VLS017', 'SO2301006', '003/2023', 'SI2301025', 'VLS035', 'D07796', 'D07840', 'D07847', 'VLS085/2023', 'D07841', '5c3c8939-0f90-d1fb-ca13-b8019f927cf9', 'P25598', 'D07871', '015/2023']

# l = ['VLS146/2023', 'VLS192/2023', 'SI2303047', 'D07984', 'VLS196/2023', 'D07982', '20230403 (CHEUNG MING)', 'D08000', 'D07999', 'D07996', '032/2023', '029/2023', 'D08024', 'D08025', 'VLS222/2023', 'P25709-R01', 'SI2304041', 'D08044', 'D08051', 'D08058', 'XK-20230426-0101', 'D08065', 'XK-20230502-0104', 'D08077', 'D08108', 'D08158', '20230515', 'VLS229/2023', '015aa66b-56ea-da29-a1ca-86d046d2a2c6']


# import pathlib

# path = r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\29-9-2023 保稅倉"

# folder = pathlib.Path(path)
# count = 0

# fileList = list(folder.iterdir())
# refDict = {}


# for file in fileList:
#     count += 1
#     ref = re.search("[A-Z][\.\d]+",file.name).group()
#     if ref in refDict:
#         refDict[ref] += 1
#     else:
#         refDict[ref] = 1

# for item in list(refDict.keys()):
#     print(item)
# # print(refDict)

# # lnot equal 3: not complete ref
# def prNotThree(ref):
#     if refDict[ref] != 3:
#         print(ref)
# print("-------------------------------------")

# print(len(list(map(prNotThree, list(refDict.keys())))))

# print(count)


# l = 1

# x = input('input a number') if l == 0 else print('hi')

# wb = openpyxl.load_workbook(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\CHARMSMART 2023 出入貨紀錄.xlsx")
# ws = wb['active']

# l = []

# for row in range(508, ws.max_row +1):
#     if ws["b" + str(row)].value != None and ws["b" + str(row)].value != 'cs' and ws["b" + str(row)].value != 'onera':
#         print(ws["b" + str(row)].value)
#         l.append(ws["b" + str(row)].value)

# 2023 invoice
# def checkInside(i):
#     for item in l:
#         if str(i).upper() in str(item).upper():
#             print(item)
#             return True
#     return False

# l = ['VLS146/2023', 'VLS192/2023', 'SI2303047', 'D07984', 'VLS196/2023', 'D07982', '20230403 (CHEUNG MING)', 'D08000', 'D07999', 'D07996', '032/2023', '029/2023', 'D08024', 'D08025', 'VLS222/2023', 'P25709-R01', 'SI2304041', 'D08044', 'D08051', 'D08058', 'XK-20230426-0101', 'D08065', 'XK-20230502-0104', 'D08077', 'D08108', 'D08158', 20230515, 'VLS229/2023', '015aa66b-56ea-da29-a1ca-86d046d2a2c6', '052/2023', 'P25688-R01', 'N000060', 20230606, 'SI2306018', 'D08218', 'SI2307018', 'D08235', 'D08236', '20230705-BC', 'f8c56d9c-8435-3c3f-c0d8-694903d7a009', 'D08242', 'D08252', 'VLS428/2023', 'VLS430/2023', 'P25897-R01', 'D08264', 'D08265', 20230727001, 'D08287', 'D08289', 'D08290', 'D08292', 'VLS453/2023', 20230728001, 'D08310', 'D08316', 'PROS230586', 'VLS482/2023', 'VLS484/2023', 20230814, 'SI2308032', 'XK-20230904-0230', 'D08401', 'D08402', 'D08403', 'D08409', '3a22307a-7c92-c4d4-aa86-7a484d37343e', '32636578-206d-6b92-114f-e2aed113f2d9', 'PROS230704', 'D08419', 'P26009', 'D08429', '085/2023', 'P25943-R01', 20230918001, 'VLS616/2023', '091/2023', 'PROS230817', '6ff575e3-0670-bcle-2fc6-9c5637481acd']
# print(checkInside('p25646'))

# def sumOfPowerTwo(n):
#     if n == 0:
#         return 1
#     else:
#         return pow(2, n) + sumOfPowerTwo(n - 1)

# print(pow(sumOfPowerTwo(7),4))

print(os.path.join())