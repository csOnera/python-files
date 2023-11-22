from selenium import webdriver
from selenium.webdriver.chrome.service import Service
# from webdriver_manager.chrome import ChromeDriverManager

from selenium.webdriver.common.by import By

from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

service = Service()
options = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=options)

# ADD ID, MONEY

import openpyxl
import re

serialPath = r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\雜\operate Serial No. by Scanning.xlsm"

wb = openpyxl.load_workbook(serialPath, read_only=False, keep_vba=True)
ws = wb.active

# insert a row first
ws.insert_cols(2)

count = 0

# def preferLongerLength(dict):
#     for key in dict:


for row in range(1, ws.max_row + 1):
    if ws["a" + str(row)].value != None:
        count += 1
        driver.get(f'https://www.google.com/search?q={ws["a" + str(row)].value}&rlz=1C1ONGR_zh-HKHK950HK951&oq={ws["a" + str(row)].value}+&gs_lcrp=EgZjaHJvbWUyBggAEEUYOdIBCTMwMTdqMGoxNagCALACAA&sourceid=chrome&ie=UTF-8')

        titles = driver.find_elements(By.XPATH, '//*[@id="rso"]/div')

        print(len(titles))

        # as a dict
        findingList = {}
        for title in titles:
            try:
                find = re.search('[A-Z]+[\d\.]+',title.text).group()
                # print(find)
                if findingList.get(find) is None and len(find) >= 5:
                    findingList[find] = 1
                else:
                    findingList[find] += 1
            except:
                # print('pattern not found')
                continue

            # print(title.find_element(By.XPATH, 'div/div/div/div[1]/div/div/span/a/h3').text)
        print(findingList)
        if findingList != {}:
            if list(findingList.keys())[0][0] == "H":
                maxChosen = max(findingList, key=findingList.get)
            else:
                maxChosen = max(findingList, key=len)
            ws["b" + str(row)].value = maxChosen


        # print(title.text)
        
    # if count == 6:
        # break
    

wb.save(serialPath)



while True:
    x = input('quit to quit')
    if x.lower() == 'quit':
        driver.quit()
        exit()