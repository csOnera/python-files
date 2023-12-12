
import openpyxl

popORpdd = input("pop or pdd?").lower()

if popORpdd == 'pdd':
    excel = openpyxl.load_workbook("D:\桌面\拼多多約倉\预约订单导入.xlsx")
    POP = openpyxl.load_workbook(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\出貨OR退貨紀錄\拼多多出貨記錄.xlsx")
else:
    excel = openpyxl.load_workbook("D:/桌面/POP 約倉/预约订单导入.xlsx")
    POP = openpyxl.load_workbook(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\出貨OR退貨紀錄\POP出貨記錄LATEST VERSION-DESKTOP-833R29B.xlsx")


sheet = excel.active
popwb = POP.active
# print(sheet["a2"].value)

for i in range(2, popwb.max_row + 1):
    if popwb["q" + str(i)].value != None:
        j = i
        l = []
        while True:
            if popwb["b" + str(j)].value != None and popwb["b" + str(j - 1)].value == None and popwb['e' + str(j-1)].value == None:
                l.append(popwb["b" + str(j)].value)
                SF = popwb["k" + str(j)].value
                if SF == None:
                    print('blank SF please re-enter later')
                boxNum = len(l)
                break
            elif popwb["b" + str(j)].value != None:
                l.append(popwb["b" + str(j)].value)
            j -= 1
        break
for i in range(len(l)):
    sheet["a" + str(i+2)].value = l[i]

if popORpdd == 'pdd':
    absFilePath = "D:\桌面\拼多多約倉\\" + str(l[0]) + " 预约订单导入.xlsx"
else:
    absFilePath = "D:\桌面\POP 約倉\\" + str(l[0]) + " 预约订单导入.xlsx"
excel.save(absFilePath)

print(SF, '\n', l)
print('入倉excel accomplished!')


# start selenium after excel done

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
# from webdriver_manager.chrome import ChromeDriverManager

from selenium.webdriver.common.by import By

from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

service = Service()
options = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=options)

driver.get("https://business.wtdex.com/login#/order/reservationordernew")

signin = driver.find_elements(By.CLASS_NAME, "form-control")

import os
from dotenv import load_dotenv

load_dotenv()

POP_UN = os.getenv('POP_UN')
POP_PW = os.getenv('POP_PW')
PDD_UN = os.getenv('PDD_UN')
PDD_PW = os.getenv('PDD_PW')


if popORpdd == "pdd":
    signin[0].send_keys(PDD_UN)
    signin[1].send_keys(PDD_PW)
else:
    signin[0].send_keys(POP_UN)
    signin[1].send_keys(POP_PW)
    
signin[2].click()

check = input("entered 驗證碼 and login? 'Y'")

if check == 'Y':
    driver.get("https://business.wtdex.com/index#/order/reservationordernew")
    
    # frame = driver.find_element(By.CSS_SELECTOR, 'div.tool_forms iframe')
    driver.switch_to.frame(1)

    element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, 'taxquota'))
    )

    
    try:
        add = driver.find_element(By.XPATH, '//*[@id="toolbar"]/a[1]')
        add.click()
        driver.switch_to.frame(0)
        if SF != None:
            driver.find_element(By.XPATH,'//*[@id="form-reservationordernew-add"]/div[1]/div/input').send_keys(SF)
        driver.find_element(By.XPATH,'//*[@id="form-reservationordernew-add"]/div[3]/div/select').send_keys('BC')
        driver.find_element(By.XPATH,'//*[@id="form-reservationordernew-add"]/div[4]/div/select').send_keys('集货')
        driver.find_element(By.XPATH, '//*[@id="form-reservationordernew-add"]/div[2]/div/select').send_keys('入仓预约')
        driver.find_element(By.XPATH,'//*[@id="form-reservationordernew-add"]/div[5]/div/input').send_keys('0')
        driver.find_element(By.XPATH,'//*[@id="form-reservationordernew-add"]/div[6]/div/input').send_keys(boxNum)
        driver.find_element(By.XPATH, '//*[@id="file"]').send_keys(absFilePath)
        import datetime
        date = str(datetime.date.today() + datetime.timedelta(days=1)) + ' 11:00'
        driver.find_element(By.XPATH,'//*[@id="inboundTime2"]').send_keys(date)
    except:
        print('error..............')
    
    finally:
        print('........DONE.......')



    while True:
        y = input('close webDriver: "close"')
        if y == 'close':
            driver.close()
            break
