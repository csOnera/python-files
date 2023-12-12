import openpyxl
import mysql.connector



import os
from dotenv import load_dotenv

load_dotenv()

MYSQL_USER = os.getenv('MYSQL_USER')
MYSQL_PW = os.getenv('MYSQL_PW')

connection = mysql.connector.connect(
    host='localhost',
    port='3306',
    user= MYSQL_USER,
    password= MYSQL_PW,
    database='trial_database'
)

cursor = connection.cursor()


try:
    excel = openpyxl.load_workbook(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\以前紀錄\exportRecords.xlsx")
    sheet = excel.active
    sheet.title = 'sheet2'
    excel.create_sheet('Sheet')
    excel.remove(sheet)
except:
    excel = openpyxl.Workbook()
sheet = excel.active


# start here
# cursor.execute("""
#     show columns from `refInvoiceNo`;
# """)
# records = cursor.fetchall()
# colNum = len(records)
# for i in records:
#     sheet.cell(row = 1, column = records.index(i) + 1).value = i[0]
l = ['id', '日期', '型號', '發票', '數量', '去處', 'price', 'ref_id']

for i in l:
    sheet.cell(row = 1, column = l.index(i)+1).value = i

cursor.execute("""
    select * from `exportRecord`;
""")
records = cursor.fetchall()
for i in range(1,len(records)+1):
    for j in range(len(l)):
        sheet.cell(row = i + 1, column = j + 1).value = records[i - 1][j]



print("succeeded!!")





excel.save(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\以前紀錄\exportRecords.xlsx")


cursor.close()
connection.commit()
connection.close()

if __name__ == "__main__":
    import win32com.client

    xl = win32com.client.Dispatch("Excel.Application")
    xl.Visible = True
    wb_open = True
    wb = xl.Workbooks.Open(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\以前紀錄\exportRecords.xlsx")
