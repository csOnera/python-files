# pre-requisite: input data in InputExport.xlsx neatly
# after inputting "Y" for confirmation, 
# this program will automatically calculate everything and 
# operate the records and numbers when
# num of stock > takingNum
# now linked with exportExport
# to be modified: no need to input num V
# hvnt tested: added ref_id



import mysql.connector
import openpyxl
from exportExport import exportExport, runVBA
import win32com.client

import os
from dotenv import load_dotenv

load_dotenv()

MYSQL_USER = os.getenv('MYSQL_USER')
MYSQL_PW = os.getenv('MYSQL_PW')

connection = mysql.connector.connect(
    host='localhost',
    port='3306',
    user= MYSQL_USER,
    password= MYSQL_PW,
    database='trial_database'
)

cursor = connection.cursor()

xl = win32com.client.Dispatch("Excel.Application")
xl.Visible = True
wb_open = True
wb = xl.Workbooks.Open(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\DATABASE_TRIAL\python files\InputExport.xlsx")



confirm = input('請檢查清楚excel表,確認請輸入"Y"\nEXCEL sheet entered? "Y" ')
autoNum = 0
priceStr = ''
lastStockL = []
errorL = []

if confirm == 'Y':
    wb.Close(False)
    try:
        eerCheck = openpyxl.load_workbook(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\DATABASE_TRIAL\python files\exportExportRecords.xlsm", read_only=False, keep_vba=True)
        eerCheck.close()

    except:
        # excel = openpyxl.Workbook()
        print("請先關閉exportExportRecords excel檔\nplease close exportExportRecords before running the app")
        import time
        time.sleep(5)
        quit()


    excel = openpyxl.load_workbook(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\DATABASE_TRIAL\python files\InputExport.xlsx")
    sheet = excel.active
    
    # autoNum = sheet.max_row - 1
    for j in range(2, sheet.max_row + 1):
        ref = sheet['a' + str(j)].value

        # break if sheet.max_row detect None rows
        if ref == None:
            print(sheet.max_row)
            break

        # check if there is this ref
        cursor.execute("select `型號` from `refInvoiceNo` where `型號` = '{}';".format(ref))
        check = cursor.fetchall()

        if check == []:
            # print("庫存沒有該型號: " + str(ref))
            errorL.append("庫存沒有該型號: " + str(ref))
            continue
        else:
            # check if enough stock
            takingNum = int(sheet['b' + str(j)].value)
            to = sheet['c' + str(j)].value
            price = sheet['d' + str(j)].value
            undefault = sheet['f' + str(j)].value

            if price == None:
                price = 0
            else:
                price = float(price)
            
            if to == None:
                print("請填寫所有空格")
                import time
                time.sleep(3)
                quit()

            # new update that add choice of selecting
            if undefault != None:
                # check if there is / if only one source
                cursor.execute("select `型號`, `發票`, `數量`, `現貨/退貨/盒` from `refInvoiceNo` where `發票` like '{}%' and `型號` = '{}' and `數量` <> 0;".format(undefault, ref))
                result = cursor.fetchall()

                if result == []:
                    errorL.append('沒在選擇的發票中找到該型號 {}'.format(ref))
                    # NOT CONTINUE
                if len(result) >= 1:
                    for item in result:
                        print(item)
                    checkInvoice = input("如發票沒有錯請填'N', 不然請填'Y': ")
                    if checkInvoice == "Y":
                        errorL.append('存在非選擇發票: {}'.format(ref))
                        # may select the id
                        continue
                    elif checkInvoice == "N":
                        # check if enough number (no need)
                        # cursor.execute("select sum(`數量`) from `refInvoiceNo` where `發票` = '{}%' and `型號` = '{}';".format(undefault, ref))
                        # totalNum = cursor.fetchall()[0][0]
                        # if takingNum <= totalNum:
                        for item in result:
                        # if first record not enough number
                            if int(item[3]) > takingNum:
                                print(1)
                                cursor.execute("update `refInvoiceNo` set `數量` = '{}' where `id` = '{}';".format(int(item[3]) - takingNum, item[0]))
                                cursor.execute("""insert into `exportRecord` (`日期`,`型號`,`發票`,`數量`, `去處`, `price`, `ref_id`) values 
                                                (curdate(), '{}', '{}', '{}', '{}', '{}', '{}');""".format(ref, item[2], takingNum, to, price, item[0]))
                                autoNum += 1
                                takingNum = 0
                                break
                            elif int(item[3]) == takingNum:
                                print(2)
                                cursor.execute("update `refInvoiceNo` set `數量` = 0 where `id` = '{}';".format(item[0]))
                                cursor.execute("""insert into `exportRecord` (`日期`,`型號`,`發票`,`數量`, `去處`, `price`, `ref_id`) values 
                                                (curdate(), '{}', '{}', '{}', '{}', '{}', '{}');""".format(ref, item[2], takingNum, to, price, item[0]))
                                autoNum += 1
                                takingNum = 0
                                break
                            # else if num < takingNum cases(if it is the last record)
                            elif result.index(item) == len(result) - 1 and int(item[3]) < takingNum:
                                print(3)
                                cursor.execute("update `refInvoiceNo` set `數量` = 0 where `id` = '{}';".format(item[0]))
                                cursor.execute("""insert into `exportRecord` (`日期`,`型號`,`發票`,`數量`, `去處`, `price`, `ref_id`) values 
                                                (curdate(), '{}', '{}', '{}', '{}', '{}', '{}');""".format(ref, item[2], item[3], to, price, item[0]))
                                autoNum += 1
                                takingNum = takingNum - int(item[3])
                                errorL.append('選擇的發票紀錄沒有足夠數量: {}'.format(ref))
                            # else if num < takingNum and not the last record
                            else:
                                print(4)
                                cursor.execute("update `refInvoiceNo` set `數量` = 0 where `id` = '{}';".format(item[0]))
                                cursor.execute("""insert into `exportRecord` (`日期`,`型號`,`發票`,`數量`, `去處`, `price`, `ref_id`) values 
                                                (curdate(), '{}', '{}', '{}', '{}', '{}', '{}');""".format(ref, item[2], item[3], to, price, item[0]))
                                autoNum += 1
                                takingNum = takingNum - int(item[3])
                                



            cursor.execute("set @ref := '{}';".format(ref))
            cursor.execute("set @takingNum := '{}';".format(takingNum))


            # if not enough stock, show total number in stock

            cursor.execute("select `數量` from `refInvoiceNo` where `型號` = '{}';".format(ref))
            noList = cursor.fetchall()

            if len(noList) > 1:
                totalNum = 0
                for i in noList:
                    totalNum += i[0]
            else:
                totalNum = noList[0][0]

            # print('totalNum: ',totalNum)
            if takingNum > totalNum:
                print('{} not enough stock (庫存不足), {} left(剩餘數量).'.format(ref, totalNum))
                if totalNum != 0:
                    stillTake = input("請問仍要輸出不足的數量嗎\nwant to export even not enough stock?'Y'")
                    if stillTake == 'Y':
                        # print(autoNum)
                        takingNum = totalNum
                        lastStockL.append(ref)
                        cursor.execute("select * from `refInvoiceNo` where `型號` = '{}';".format(ref))
                        refInfo = cursor.fetchall()
                        for i in refInfo:
                            print(i)
                        while takingNum > 0:
                            # cursor.execute("set @num := (select `數量` from `refInvoiceNo` where `數量` <> 0 and `型號` = '{}' order by `id` asc limit 1);".format(ref))
                            # cursor.execute("set @id := (select `id` from `refInvoiceNo` where `數量` <> 0 and `型號` = '{}' order by `id` asc limit 1);".format(ref))
                            
                            
                            cursor.execute("""
                                set @id = (select ifnull((select `id` from `refInvoiceNo` 
                                where `數量` <> 0 
                                and `型號` = '{}'
                                and `現貨/退貨/盒` = '退貨'
                                order by `id` asc limit 1),
                                (select `id` from `refInvoiceNo`
                                where `數量` <> 0 
                                and `型號` = '{}'
                                order by `id` asc limit 1)));
                            """.format(ref,ref))
                            cursor.execute("""
                                set @num = (select `數量` 
                                from `refInvoiceNo` where
                                `id` = @id); 
                            """)


                            cursor.execute("select @num;")
                            num = cursor.fetchall()[0][0]

                            cursor.execute("select `cost` from `refInvoiceNo` where `id` = @id;")
                            priceStr += '\n' + str(cursor.fetchall()[0][0])

                            if takingNum < num:
                                #  minus number in refInvoiceNo table
                                cursor.execute("update `refInvoiceNo` set `數量` = '{}' where `id` = @id;".format(num - takingNum))
                                #  record in exportRecord table
                                cursor.execute("insert into `exportRecord` (`日期`,`型號`,`發票`,`數量`, `去處`, `price`, `ref_id`) values (curdate(), @ref, (select `發票` from `refInvoiceNo` where `id` = @id), '{}', '{}', '{}', @id);".format(takingNum,to,price))
                                autoNum += 1
                                takingNum = 0
                            else:
                                cursor.execute("insert into `exportRecord` (`日期`,`型號`,`發票`,`數量`, `去處`, `price`, `ref_id`) values (curdate(), @ref, (select `發票` from `refInvoiceNo` where `id` = @id), (select `數量` from `refInvoiceNo` where `id` = @id), '{}','{}', @id);".format(to,price))
                                cursor.execute("update `refInvoiceNo` set `數量` = 0 where `id` = @id;")
                                autoNum += 1
                                takingNum -= num
                        print('---------------------------------------------------------------')
                else:
                    errorL.append('{} not enough stock.'.format(ref))


            else:
                if takingNum == totalNum:
                    lastStockL.append(ref)
                cursor.execute("select * from `refInvoiceNo` where `型號` = '{}';".format(ref))
                refInfo = cursor.fetchall()
                for i in refInfo:
                    print(i)
                while takingNum > 0:
                    # cursor.execute("set @num := (select `數量` from `refInvoiceNo` where `數量` <> 0 and `型號` = '{}' order by `id` asc limit 1);".format(ref))
                    # cursor.execute("set @id := (select `id` from `refInvoiceNo` where `數量` <> 0 and `型號` = '{}' order by `id` asc limit 1);".format(ref))
                    cursor.execute("""
                        set @id = (select ifnull((select `id` from `refInvoiceNo` 
                        where `數量` <> 0 
                        and `型號` = '{}'
                        and `現貨/退貨/盒` = '退貨'
                        order by `id` asc limit 1),
                        (select `id` from `refInvoiceNo`
                        where `數量` <> 0 
                        and `型號` = '{}'
                        order by `id` asc limit 1)));
                    """.format(ref,ref))
                    cursor.execute("""
                        set @num = (select `數量` 
                        from `refInvoiceNo` where
                        `id` = @id); 
                    """)
                    
                    cursor.execute("select @num;")
                    num = cursor.fetchall()[0][0]

                    cursor.execute("select `cost` from `refInvoiceNo` where `id` = @id;")
                    priceStr += '\n' + str(cursor.fetchall()[0][0])

                    if takingNum < num:
                        #  minus number in refInvoiceNo table
                        cursor.execute("update `refInvoiceNo` set `數量` = '{}' where `id` = @id;".format(num - takingNum))
                        #  record in exportRecord table
                        cursor.execute("insert into `exportRecord` (`日期`,`型號`,`發票`,`數量`, `去處`, `price`, `ref_id`) values (curdate(), @ref, (select `發票` from `refInvoiceNo` where `id` = @id), '{}', '{}', '{}', @id);".format(takingNum,to,price))
                        autoNum += 1
                        # print(autoNum)
                        takingNum = 0
                    else:
                        cursor.execute("insert into `exportRecord` (`日期`,`型號`,`發票`,`數量`, `去處`, `price`, `ref_id`) values (curdate(), @ref, (select `發票` from `refInvoiceNo` where `id` = @id), (select `數量` from `refInvoiceNo` where `id` = @id), '{}','{}', @id);".format(to,price))
                        cursor.execute("update `refInvoiceNo` set `數量` = 0 where `id` = @id;")
                        autoNum += 1
                        # print(autoNum)
                        takingNum -= num
                print('---------------------------------------------------------------')
else:
    exit()

cursor.close()
connection.commit()
connection.close()

print("==================================================================")
# print(autoNum, priceStr)
for i in lastStockL:
    print(i + "庫存零\n")

for i in errorL:
    print(i)

print("==================================================================")

print(autoNum)

exportExport("num",autoNum)

# here run vba/other to input cost and invoice to POP

# CHECK IF POP OPENED??


# pop = xl.Workbooks.Open(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\出貨OR退貨紀錄\POP出貨記錄LATEST VERSION-DESKTOP-833R29B.xlsx")
wb = xl.Workbooks.Open(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\DATABASE_TRIAL\python files\exportExportRecords.xlsm")

xl.Application.Run('exportExportRecords.xlsm!Module3.autoAddCostNOriginToPOP')
# pop.Close(True)

runVBA()

import RefreshStock
RefreshStock
import RefreshRecord
RefreshRecord

while True:
    quitT = input('可輸入"quit"來退出或按視窗的"X"\ntype "quit" to quit terminal')
    if quitT == "quit":
        quit()