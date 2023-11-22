# 已補C0 & CS2023 (5/6/2023)
# 

import openpyxl
import re

invoiceC = openpyxl.load_workbook(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\Invoice C.xlsx")
cs2023 = openpyxl.load_workbook(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\CHARMSMART 2023 出入貨紀錄.xlsx")

c = invoiceC.active
cs = cs2023.active

st = "C0\d\d"

def findRef(row):
    while True:
        if cs["d" + str(row)].value == None:
            row -= 1
        else:
            return row

def findNum(row):
    count = c["e" + str(row)].value
    while True:
        if c["e" + str(row + 1)].value != None and c["d" + str(row + 1)].value == None:
            count += c["e" + str(row + 1)].value
        else:
            return count
        row += 1



for i in range(2, cs.max_row + 1):
    # print(type(cs["k" + str(i)].value))
    if type(cs["K" + str(i)].value) == str:
        c0 = re.findall(st, cs["K" + str(i)].value)
        if c0 != []:
            row = findRef(i)
            ref = cs["d" + str(row)].value
            for j in range(2, c.max_row):
                if c["d" + str(j)].value == ref and c["l" + str(j)].value == c0[0]:
                    # print(ref,c0,findNum(j),c["o" + str(j)].value/findNum(j))
                    cs["m" + str(i)].value = c["o" + str(j)].value/findNum(j)
            



invoiceC.save(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\Invoice C.xlsx")
cs2023.save(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\CHARMSMART 2023 出入貨紀錄.xlsx")
