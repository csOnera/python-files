
import mysql.connector
import openpyxl
import win32com.client


import os
from dotenv import load_dotenv

load_dotenv()

MYSQL_USER = os.getenv('MYSQL_USER')
MYSQL_PW = os.getenv('MYSQL_PW')

connection = mysql.connector.connect(
    host='localhost',
    port='3306',
    user= MYSQL_USER,
    password= MYSQL_PW,
    database='trial_database'
)

cursor = connection.cursor()

# ask all belonging columns 
print("請輸入表格中項目的列\ne.g. A/B/C...:  ")
引號col = input("引號col: ")
型號col = input("型號col: ")
CAP號col = input("CAP號col: ")
售價HKDcol = input("售價HKDcol: ")
對賬單號col = input("對賬單號col: ")

# here enter the loading sheet
wb_path = r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\得物對賬//" + input("請關閉EXCEL後 輸入 交易成功 EXCEL名稱\ne.g. xxx.xlsx : ")
try:
    excel = openpyxl.load_workbook(wb_path)
except:
    print("請先關閉相關excel文檔")
    check = input("關閉相關excel文檔後輸入'Y'來繼續:  ")
    if check == 'Y':
        # continue
        excel = openpyxl.load_workbook(wb_path)
        
ws = excel.active

for i in range(2, ws.max_row):
    if ws["c" + str(i)].value != None and ws["w" + str(i)].value != None:
        # bug by 引號 lol
        # 引號 the primary key (so prob will get bug when inserting the same item)
        
        引號 = ws[引號col + str(i)].value[1:]
        型號 = ws[型號col + str(i)].value
        CAP號 = ws[CAP號col + str(i)].value
        售價HKD = float(ws[售價HKDcol + str(i)].value)
        對賬單號 = ws[對賬單號col + str(i)].value

        # print("""insert into `得賬` (`引號`, `型號`, `CAP號`, `售價HKD`, `對賬單號`) values 
        #       ('{}', '{}', '{}', {});"""
        #     .format(
        #         引號, 型號, CAP號, 售價HKD
        #     ))

        # added description for forseeable error of same 業務單號
        try:
            cursor.execute("""insert into `得賬` (`引號`, `型號`, `CAP號`, `售價HKD`, `對賬單號`) 
                       VALUES ('{}', '{}', '{}', {}, '{}');
                       """.format(
                           引號, 型號, CAP號, 售價HKD, 對賬單號
                       ))
            print("成功輸入單號資料: " + str(引號))
        except:
            print('error: 可能存在重覆業務單號: ' + str(引號))
            continue
print("data added")

excel.save(wb_path)

cursor.close()
connection.commit()
connection.close()

while True:
    pass