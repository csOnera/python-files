import openpyxl
import mysql.connector



connection = mysql.connector.connect(
    host='localhost',
    port='3306',
    user='root',
    password='jdysz',
    database='trial_database'
)

cursor = connection.cursor()


try:
    excel = openpyxl.load_workbook(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\庫存exportedStocks.xlsx")
    sheet = excel.active
    sheet.title = 'sheet2'
    excel.create_sheet('Sheet')
    excel.remove(sheet)
except:
    excel = openpyxl.Workbook()
sheet = excel.active


# start here
# cursor.execute("""
#     show columns from `refInvoiceNo`;
# """)
# records = cursor.fetchall()
# colNum = len(records)
# for i in records:
#     sheet.cell(row = 1, column = records.index(i) + 1).value = i[0]
l = ['型號', '數量', '發票', '現貨/退貨/盒', 'cost']

for i in l:
    sheet.cell(row = 1, column = l.index(i)+1).value = i

cursor.execute("""
    select `型號`, `數量`, `發票`, `現貨/退貨/盒`, `cost` from `refInvoiceNo`
    where `數量` <> 0 order by `型號`;
""")
records = cursor.fetchall()
for i in range(1,len(records)+1):
    for j in range(len(l)):
        sheet.cell(row = i + 1, column = j + 1).value = records[i - 1][j]



print("succeeded!!")





excel.save(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\庫存exportedStocks.xlsx")


cursor.close()
connection.commit()
connection.close()

if __name__ == "__main__":
    import win32com.client

    xl = win32com.client.Dispatch("Excel.Application")
    xl.Visible = True
    wb_open = True
    wb = xl.Workbooks.Open(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\庫存exportedStocks.xlsx")
