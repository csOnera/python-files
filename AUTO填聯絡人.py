from selenium import webdriver
from selenium.webdriver.chrome.service import Service
# from selenium.webdriver.chrome.service import Service
# from webdriver_manager.chrome import ChromeDriverManager

from selenium.webdriver.common.by import By

from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

# ADD ID, MONEY
# service_obj = Service("chromedriver.exe")
# driver = webdriver.Chrome(service=service_obj)
# driver.get("https://porder.shop.jd.com/order/orderlist/waitInnerShip?t=1687253860859")




import openpyxl



POP = openpyxl.load_workbook(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\出貨OR退貨紀錄\POP出貨記錄LATEST VERSION-DESKTOP-833R29B.xlsx")
popwb = POP.active
# print(sheet["a2"].value)

for i in range(1500, popwb.max_row + 1):
    # find the stopper
    if popwb["q" + str(i)].value != None:
        j = i
        l = []
        while True:
            if popwb["b" + str(j)].value != None and popwb["b" + str(j - 1)].value == None and popwb['e' + str(j-1)].value == None:
                l.append(popwb["b" + str(j)].value)
                break
            elif popwb["b" + str(j)].value != None:
                l.append(popwb["b" + str(j)].value)
            j -= 1
        break
print(l)
l.reverse()

lToStr = str(l[0])
for i in range(1,len(l)):
    lToStr += ',' + str(l[i])


# FIND NAME AND ADDRESS BY PRINTING PAGE
# service_obj = Service("chromedriver.exe")

service = Service()
options = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=options)

import os
from dotenv import load_dotenv

load_dotenv()

打單UN = os.getenv('PRINT_UN')
打單PW = os.getenv('PRINT_PW')

driver.get('http://amc.wtdex.com/amc/login;jsessionid=D7DAF284BAA77F5858E9945D3FC2C1CA')

# try adding a webDriverWait
try:
    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '/html/body/div/form/div/div[2]/div/div[2]/div/div[1]/span/input'))
    )
except:
    print('no response in 10 seconds, plz try again later')
    time.sleep(10)
    quit()
driver.find_element(By.XPATH,'/html/body/div/form/div/div[2]/div/div[2]/div/div[1]/span/input').send_keys(打單UN)
driver.find_element(By.XPATH, '/html/body/div/form/div/div[2]/div/div[2]/div/div[2]/span/input').click()
driver.find_element(By.XPATH, '/html/body/div/form/div/div[2]/div/div[2]/div/div[2]/span/input[2]').send_keys(打單PW)
driver.find_element(By.XPATH, '//*[@id="VCode"]').click()
# time.sleep(2)

# from PIL import Image
# import tesserocr
# import requests
# from io import BytesIO
# for i in range(4):
#     driver.find_element(By.XPATH, '//*[@id="VCode"]').click()
#     image = driver.find_element(By.XPATH, '//*[@id="VCode"]')
#     src = image.get_attribute('src')
    
#     response = requests.get(src)
#     img = Image.open(BytesIO(response.content))

#     img = img.convert('L')
#     #这个是二值化阈值
#     threshold = 150   
#     table = []

#     for i in  range(256):
#         if i < threshold:
#             table.append(0)
#         else:
#             table.append(1)
#     #通过表格转换成二进制图片，1的作用是白色，不然就全部黑色了
#     img = img.point(table,"1")
#     img.show()

#     result = tesserocr.image_to_text(img)
#     print(result)

# driver.find_element(By.XPATH, '/html/body/div/form/div/div[2]/div/div[2]/div/div[3]/span/input').send_keys(result)


driver.find_element(By.XPATH, '/html/body/div/form/div/div[2]/div/div[2]/div/div[3]/span/input').click()

check = input("entered 驗證碼 and login? 'Y'")

if check == 'Y':
    driver.get('http://amc.wtdex.com/print/packaging/toOrderPrintBatch')
    driver.find_element(By.XPATH, '//*[@id="ids"]').send_keys(lToStr)
    driver.find_element(By.XPATH, '//*[@id="query"]').click()
    # driver.find_element(By.XPATH, '//*[@id="okandno"]/div[1]/div/a[3]').click()
    test = driver.find_element(By.CLASS_NAME, 'l-grid-body-table')
    
    strList = test.text.split('\n')

    # try to get the phone number as well
    import re
    mysource = driver.page_source
    phoneList = re.findall("'recphone':'\d+'", mysource)
    realPhoneList = []
    for i in range(len(l)):
        realPhoneList.append(re.search('\d+', phoneList[i]).group())

    print(realPhoneList)



    # for i in range(len(l)):
    #     print(strList[3 + (i-1)*(5)],strList[4 + (i-1)*(5)])
    POP = openpyxl.load_workbook(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\出貨OR退貨紀錄\POP出貨記錄LATEST VERSION-DESKTOP-833R29B.xlsx")
    popwb = POP.active
    phoneIndex = 0
    for j in range(1000, popwb.max_row):
        if popwb['b' + str(j)].value in l and popwb['h' + str(j)].value == None and popwb['i' + str(j)].value == None:
            popwb['h' + str(j)].value = strList[3 + (l.index(popwb['b' + str(j)].value))*(5)]
            print(strList[3 + (l.index(popwb['b' + str(j)].value))*(5)], end="  ")
            popwb['i' + str(j)].value = strList[4 + (l.index(popwb['b' + str(j)].value))*(5)]
            print(strList[4 + (l.index(popwb['b' + str(j)].value))*(5)], end="  ")
            popwb['j' + str(j)].value = realPhoneList[phoneIndex]
            print(realPhoneList[phoneIndex])
            phoneIndex += 1
            
    POP.save(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\出貨OR退貨紀錄\POP出貨記錄LATEST VERSION-DESKTOP-833R29B.xlsx")

    driver.find_element(By.XPATH, '//*[@id="okandno"]/div[1]/div/a[1]').click()
    driver.find_element(By.XPATH, '//*[@id="okButtonDiv"]/a[1]').click()
    # driver.find_element(By.XPATH, '')
    

    while True:
        y = input('close webDriver: "close"')
        if y == 'close':
            driver.close()
            break
quit()