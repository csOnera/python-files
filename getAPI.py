import requests
import json
import time
import re

import openpyxl

serialPath = r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\雜\operate Serial No. by Scanning.xlsm"

wb = openpyxl.load_workbook(serialPath, read_only=False, keep_vba=True)
ws = wb.active

count = 0
for row in range(1,5):
    if count == 5:
        # time.sleep(60)
        count = 0
    if ws["a" + str(row)].value != None:

        api = requests.get('https://api.upcitemdb.com/prod/trial/lookup?upc={}'.format(str(ws["a" + str(row)].value)))
        count += 1
        data = api.json()

        if data.get('items') != None:
            refNum = data.get('items')[0].get('title')
            ws["b" + str(row)].value = data.get('items')[0].get('model')
            # refNum = re.search("[a-zA-Z.\d\-]+\Z", refNum).group()
            # print(refNum, data.get('items')[0].get('title'))

            # ws["c" + str(row)].value = refNum
        else:
            print(str(ws["a" + str(row)].value) + "not found")
            print(data)

wb.save(serialPath)
