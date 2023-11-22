import re
import openpyxl

cs = openpyxl.load_workbook(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\以前紀錄\CHARMSMART 2022 財政年度 出·入貨表ver2.xlsx")
csws = cs['operation page']

# outputExcel = openpyxl.load_workbook(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\2022 從出入庫存.xlsx")
# outws = outputExcel.active

# outputExcel.save(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\2022 從出入庫存.xlsx")

tempValue = ""
tempSum = 0
repeatStartingRow = 1

count = 0

for row in range(2,csws.max_row + 1):
    magicToken = 0
    # select the one can know the export number
    if csws['j' + str(row)].value != None and csws['k' + str(row)].value != None:
        if re.search("\d+-*\d+", str(csws['k'+ str(row)].value)) != None:
            if len(re.search("\d+-*\d+", str(csws['k'+ str(row)].value)).group()) >= 12 or re.search('AP', str(csws['k'+ str(row)].value)) != None:
                tempValue = csws['j' + str(row)].value
                magicToken = 1
            else:
                tempValue = csws['k' + str(row)].value
                magicToken = 1
        else:
            tempValue = csws['k' + str(row)].value
            magicToken = 1
    elif csws['j' + str(row)].value != None:
        tempValue = csws['j' + str(row)].value
        magicToken = 1
    # print(tempValue)
    # number (bug: included all possible rows)
    if magicToken == 1:
        if re.search('-\d+', str(tempValue)) != None:
            # check locus if same ref number
            if csws['d' + str(row -1)].value == csws['d' + str(row)].value and csws['e' + str(row -1)].value == csws['e' + str(row)].value:
                tempSum += int(re.search('-\d+', tempValue).group()[1:])
            else:
                tempSum = int(re.search('-\d+', tempValue).group()[1:])
        else:
            tempSum = int(csws['e' + str(row)].value)
        
        if tempSum > int(csws['e' + str(row)].value):
            print(row, tempSum, int(csws['e' + str(row)].value))
            count += 1
print(count)
    


#     # print(re.search(' -\d+', str(csws['j' + str(row)].value)).group()[2:])
#     # print(str(csws['j' + str(row)].value))
#     if re.search(' -\d+', str(csws['j' + str(row)].value)).group()[2:] == str(csws['e' + str(row)].value):
#         print(row)

