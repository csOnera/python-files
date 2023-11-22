import openpyxl
import re

cs = openpyxl.load_workbook(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\批發做單\出入貨單\ONERA 出，入貨 2022報稅.xlsx")
csWs = cs['operation page']


count = 1
invoiceList = []
innerList = []
# find invoice-------------------------
for i in range(2, csWs.max_row + 1):
    if csWs['b'+ str(i)].value != None and re.search('23-+', str(csWs['b'+ str(i)].value)) == None:
        if count == 3 or i == 1940:
            invoiceList.append(innerList)
            innerList = [{csWs['b'+str(i)].value: str(csWs['a'+str(i)].value)[:10]}]
            count = 1
        else:
            innerList.append({csWs['b'+str(i)].value: str(csWs['a'+str(i)].value)[:10]})
            count += 1


# find non B0 and C0
def locateInvoice(invoice):
    for i in range(2, csWs.max_row + 1):
        if csWs['b'+str(i)].value == invoice:
            return i


def findUpNDown(i):
    l = invoiceList[i]
    if i == len(invoiceList) - 1:
        return [locateInvoice(list(invoiceList[i][0])[0]), csWs.max_row + 1]
    else:
        return [locateInvoice(list(invoiceList[i][0])[0]), locateInvoice(list(invoiceList[i + 1][0])[0])]


def giveRowInfo(row, jork):
    ref = csWs['d'+ str(row)].value
    cost = csWs['h'+ str(row)].value
    if re.search('-\d+', str(csWs[jork+ str(row)].value)) == None:
        num =  csWs['e'+ str(row)].value
    else:
        num = re.search('-\d+', csWs[jork+ str(row)].value).group()[1:]

    return [ref,cost,num]

def invoicemaker(num):
    n = len(str(num))
    return "CS" + "0"*(3-n) + str(num)



# print(list(invoiceList[0][0])[0])
print(invoiceList)

totalNum = 0
totalPrice = 0
ref = 0
count = 1
# looping for each three-groups
for k in range(len(invoiceList)):
    [upper, lower] = findUpNDown(k)
    makeUpInvoice = invoicemaker(count)
    count += 1
    # print(list(invoiceList[k][1].values())[0])
    makeUpDate = list(invoiceList[k][len(invoiceList[k])-1].values())[0]
    infoList = []

    # operation below (i is the row number)
    for i in range(upper, lower):
        # here if statement checking the real counting receipts
        if csWs['j'+ str(i)].value != None and csWs['k'+ str(i)].value != None:
            # below changed
            if re.search('[B]0\d+',str(csWs['k'+ str(i)].value)) != None:
                # count += 1
                # if len(re.search("\d+-*\d+", str(csWs['k'+ str(i)].value)).group()) >= 12 and re.search('AP', str(csWs['k'+ str(i)].value)) == None:
                #     print(csWs['j'+ str(i)].value)
                #     [ref, cost, num] = giveRowInfo(i,'j')
                # else:
                #     print(csWs['k'+ str(i)].value)
                #     [ref, cost, num] = giveRowInfo(i,'k')
                print(csWs['k'+ str(i)].value)
                [ref, cost, num] = giveRowInfo(i,'k')
            else:
                ref = 0
        elif csWs['j'+ str(i)].value != None and str(csWs['j'+ str(i)].value).lower() != 'stock':
            # below changed
            if re.search('[B]0\d+', str(csWs['j'+ str(i)].value)) != None:
                # count += 1
                print(csWs['j'+ str(i)].value)
                [ref, cost, num] = giveRowInfo(i,'j')
            else: ref = 0
        else:
            ref = 0
        if ref != 0:
            print([ref,cost,num])
            infoList.append([ref, cost, num])

    template = openpyxl.load_workbook(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\批發做單\oneraInvoice template.xlsx")
    tempWs = template.active
    # print(tempWs['h9'].value)

    tempWs.insert_rows(14, len(infoList))
    tempWs['g7'].value = makeUpInvoice
    tempWs['h9'].value = makeUpDate
    print(makeUpInvoice)
    for row in range(13, 13 + len(infoList)):
        tempWs['b' + str(row)].value = infoList[row - 13][0]
        tempWs['e' + str(row)].value = infoList[row - 13][2]
        tempWs['g' + str(row)].value = infoList[row - 13][1]
        tempWs['h' + str(row)].value = "=g{}*e{}".format(row,row)
        # print(infoList[row - 13][2], int(infoList[row - 13][1]))
        totalPrice += int(infoList[row - 13][2]) * int(infoList[row - 13][1])
        totalNum += int(infoList[row - 13][2])
        if infoList[row - 13][0][:1] == "T":
            tempWs['a' + str(row)].value = "TISSOT"
        elif infoList[row - 13][0][:1] == "M":
            tempWs['a' + str(row)].value = "MIDO"
        elif infoList[row - 13][0][:1] == "L":
            tempWs['a' + str(row)].value = "LONGINE"
        elif infoList[row - 13][0][:1] == "C":
            tempWs['a' + str(row)].value = "CERTINA"
        else:
            tempWs['a' + str(row)].value = "SWATCH"
    tempWs['e' + str(15 + len(infoList))].value = "=sum(e13:e{})".format(14 + len(infoList))
    tempWs['h' + str(15 + len(infoList))].value = "=sum(h13:h{})".format(14 + len(infoList))
    tempWs['g' + str(16 + len(infoList))].value = 0.003
    tempWs['h' + str(16 + len(infoList))].value = "=1.003*h{}".format(15+len(infoList))
    bineName = str(list(invoiceList[k][0])[0])
    for i in range(1,len(invoiceList[k])):
        bineName += "/" + str(list(invoiceList[k][i])[0])

    tempWs['b11'].value = bineName





    template.save(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\批發做單\onera to cs invoiceAgain\\" + makeUpInvoice + '.xlsx')

    # break





# print(count)

print(totalPrice, totalNum)


cs.close()