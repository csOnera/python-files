import os
import openpyxl
import pathlib
import re
import time

folderName = input("please input the folder name\nplease make sure it is inside '得物對賬': ")

folderPath =  r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\得物對賬/" + folderName



folder = pathlib.Path(folderPath)
folderL = list(folder.iterdir())

import warnings

for file in folderL:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(file)
        ws = wb['账单总览']
        name = ws["a3"].value + ".xlsx"

        # 下加引號
        ws = wb['鉴定通过订单']
        for row in range(3,ws.max_row+1):
            if ws["a" + str(row)].value != None:
                ws["b" + str(row)].value = "'" + str(ws["a" + str(row)].value)
        print(os.path.join(str(folder), name))
        wb.save(os.path.join(str(folder), name))
        wb.close()
