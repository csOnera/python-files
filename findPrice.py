from selenium import webdriver
from selenium.webdriver.chrome.service import Service
# from webdriver_manager.chrome import ChromeDriverManager

from selenium.webdriver.common.by import By

from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

service = Service()
options = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=options)

# ADD ID, MONEY

import openpyxl
import re

serialPath = r"D:\Downloads\複本 汉米新品 (24.3.22提报).xlsx"

wb = openpyxl.load_workbook(serialPath, read_only=False, keep_vba=True)
ws = wb.active

# insert a row first
# ws.insert_cols(2)

count = 0

# def preferLongerLength(dict):
#     for key in dict:


for row in range(2, ws.max_row + 1):
    if ws["a" + str(row)].value != None:
        count += 1
        try:
            ref = re.search('[A-Z\d\.\-]+$',ws["a" + str(row)].value).group()
        except:
            print('error: ',ws["a" + str(row)].value)    
        driver.get(f'https://www.hamiltonwatch.com/en-hk/catalogsearch/result/?q={ref}')

        price = driver.find_elements(By.XPATH, '//*[@id="maincontent"]/div[3]/div[1]/div/div[2]/ol/li/div/div[2]/div[3]')

        existCheck = driver.find_elements(By.XPATH, '//*[@id="maincontent"]/div[3]/div[1]/div/dl/dt')
        

        print(ref)
        print(price[0].text)

        try:
            print(existCheck[0].text)
        except:
            pass




        # findingList = {}
        # for title in titles:
        #     # print(title.text)
        #     try:
        #         find = re.search('\d+\sg',title.text).group()
                
        #         if findingList.get(find) is None:
        #             findingList[find] = 1
        #         else:
        #             findingList[find] += 1
        #     except:
        #         # print('pattern not found')
        #         continue

        #     # print(title.find_element(By.XPATH, 'div/div/div/div[1]/div/div/span/a/h3').text)
        # print(findingList)
        # if findingList != {}:
        #     if list(findingList.keys())[0][0] == "H":
        #         maxChosen = max(findingList, key=findingList.get)
        #     else:
        #         maxChosen = max(findingList, key=len)
        #     ws["b" + str(row)].value = maxChosen


        # print(title.text)
        
    # if count == 6:
        # break
    

# wb.save(serialPath)



while True:
    x = input('quit to quit')
    if x.lower() == 'quit':
        driver.quit()
        exit()