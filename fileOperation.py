from datetime import datetime
import time

thisYear = datetime.now().year #integer


def create_file(filename=r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\DATABASE_TRIAL\python files\year.txt"):
    try:
        with open(filename, 'w') as f:
            f.write(str(thisYear))
        print("File " + filename + " created successfully with year.")
    except IOError:
        print("Error: could not create file " + filename)
 
def read_file(filename= r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\DATABASE_TRIAL\python files\year.txt"):
    try:
        with open(filename, 'r') as f:
            year = f.read()
            return year
    except IOError:
        print("Error: could not read file " + filename)
 
def append_file(filename, text):
    try:
        with open(filename, 'a') as f:
            f.write(text)
        print("Text appended to file " + filename + " successfully.")
    except IOError:
        print("Error: could not append to file " + filename)


if __name__ == '__main__':
    try:
        year = read_file()
    except:
        create_file("year.txt")
        print(f"Reminder: the newly created file is with year {thisYear}, if now is {thisYear} but before 1st of April, please edit the year manually back to {thisYear - 1}")

    print(f"Current year in record is {year if year != '' else 'not found'}")

    while True:
        answer = input(f"input 'change year' to turn the year to {thisYear}:  ")

        if answer == 'change year':
            create_file()
            # print(f"successfully changed year to {thisYear}")
            # time.sleep(10)

            # create new excel with sheet "舊貨庫存"

            import openpyxl
            newFilePath = r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\CHARMSMART "+ str(thisYear) +" 出入貨紀錄.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            firstLine = ['日期','發票','品牌','型號','數量','公價','折扣','單價','金額','客戶','發票','折扣 (折)','金額','利潤','北京','外幣','運費/加運費後單價','JOUSTHN 雜費計算']
            ws.title = 'active'

            for i in range(len(firstLine)):
                # check if below works 
                ws.cell(row=1, column=i+1).value = firstLine[i]

            ws1 = wb.create_sheet('舊貨庫存')

            # below import stock from local mysql as Old Stock
            import os
            from dotenv import load_dotenv
            import mysql.connector

            load_dotenv()

            MYSQL_USER = os.getenv('MYSQL_USER')
            MYSQL_PW = os.getenv('MYSQL_PW')

            connection = mysql.connector.connect(
                host='localhost',
                port='3306',
                user= MYSQL_USER,
                password= MYSQL_PW,
                database='trial_database'
            )

            cursor = connection.cursor()

            oldStockFirstLine = ['id', '發票','現貨/退貨/盒','型號','數量','cost', 'cs/onera', '入貨日期']
            
            for i in range(len(oldStockFirstLine)):
                # check if below works 
                ws1.cell(row=2, column=i+1).value = oldStockFirstLine[i]
            ws1.cell(row=2, column=10).value = '發貨'
            ws1.cell(row=2, column=12).value = '現數量'
            ws1.cell(row=2, column=13).value = 'price'

            cursor.execute("""
                select `id`, `發票`,`現貨/退貨/盒`,`型號`,`數量`,`cost`, `cs/onera`, `入貨日期` from refInvoiceNo
                where `數量` <> 0
                order by `發票`;
            """)
            result = cursor.fetchall()
            for i in range(len(result)):
                ws1["l" + str(i + 3)].value = result[i][4]
            
            for i in range(len(result)):
                for j in range(len(oldStockFirstLine)):
                    ws1.cell(row = i + 3, column = j + 1).value = result[i][j]

            wb.save(newFilePath)

            print('succeeded')
            time.sleep(10)
            exit()
        else:
            print("please input valid answer")
            time.sleep(10)

