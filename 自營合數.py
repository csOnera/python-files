import pathlib
import os
import openpyxl
import re

folderPath = r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\報關單2022"
folder = pathlib.Path(folderPath)

flist = list(folder.iterdir())

storageList = []

for file in flist:
    if os.path.isfile(file) and file.name != "總數.xlsx":
        wb = openpyxl.load_workbook(folderPath + "\\" + file.name)
        ws = wb['箱单']
        for row in range(5,ws.max_row):
            if type(ws['d' + str(row)].value) == str:
                if re.search('[a-zA-Z]+[\d.]+', ws['d' + str(row)].value) != None:
                    ref = re.search('[a-zA-Z]+[\d.]+', ws['d' + str(row)].value).group()
                    if ref == "SO27" or ref == "PH500" or ref == "PH200":
                        print(file.name)
                        print(ref, ws['e' + str(row)].value)
                    storageList.append({ref : ws['e' + str(row)].value})
        wb.close()


# exportfile = openpyxl.load_workbook(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\報關單2022\總數.xlsx")
# exportfilews = exportfile.active

# for row in range(1, len(storageList) + 1):
#     exportfilews['a' + str(row)].value = list(storageList[row - 1].keys())[0]
#     exportfilews['b' + str(row)].value = list(storageList[row - 1].values())[0]

# exportfile.save(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\報關單2022\總數.xlsx")