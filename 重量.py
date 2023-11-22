import openpyxl

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


service = Service()
options = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=options)

excel = openpyxl.load_workbook(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\報關單\資料處理檔.xlsm", read_only=False, keep_vba=True)
ws = excel.active

weightExcel = openpyxl.load_workbook(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\雜\過往型號重量.xlsx")
wws = weightExcel['no repeat']

priceList = []

def findFromThePast(ref):
    for row in range(1,wws.max_row+1):
        if wws["a" + str(row)].value == ref:
            return wws["b" + str(row)].value
    return None


for row in range(2,4):
    ref = ws["f" + str(row)].value
    if ref != None:
        weight = findFromThePast(ref)
        if weight != None:
            priceList.append(weight)
            continue
        match ref[:1].lower():
            case "t":
                driver.get("https://www.tissotwatches.com/en-hk/catalogsearch/result/?q=" + ref)
                driver.find_element(By.XPATH, '//*[@id="onetrust-accept-btn-handler"]').click()
                WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="tissot-details-1"]'))).click()
                try:
                    weight = driver.find_element(By.XPATH,'//*[@id="tissot-details-1-content"]/div/ul/li[2]/span[2]').get_attribute('textContent')
                    priceList.append(weight)
                except:
                    print(f"{ref} weight not found")            
            case "m":
                driver.get("www.midowatches.com/en/catalogsearch/result/?q=" + ref)
            case "l":

                # cant find universal query website??
                pass
            case default:
                # find in a table (excel or database)
                pass
print(priceList)
