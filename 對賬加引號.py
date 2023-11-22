import openpyxl
import pathlib

def addAllSheets():
    # CHANGE THE PATH BELOW
    excelPath = r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\得物對賬\(勝)04-2023至03-2024年 得物對帳單整合.xlsx"
    excel = openpyxl.load_workbook(excelPath)

    for i in range(len(excel.sheetnames)):
        excel.active = i
        ws = excel.active
        
        for row in range(2,ws.max_row+1):
            if ws["a" + str(row)].value != None:
                ws["b" + str(row)].value = "'" + str(ws["a" + str(row)].value)
    excel.save(excelPath)

def addOneFolder():
    # CHANGE THE PATH BELOW
    folder = pathlib.Path(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\得物對賬\23 6.01-7.31")
    flist = list(folder.iterdir())
    for file in flist:
        excel = openpyxl.load_workbook(file)
        ws = excel['鉴定通过订单']
        for row in range(3,ws.max_row+1):
            if ws["a" + str(row)].value != None:
                ws["b" + str(row)].value = "'" + str(ws["a" + str(row)].value)
                # print(ws["a" + str(row)].value)
        excel.save(file)

# addOneFolder()
# addAllSheets()