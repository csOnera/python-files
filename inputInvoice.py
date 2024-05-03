# pre-requisite: data inputted in InputInvoice.xlsx
# this program will add the data of invoice into RefInvoiceNo table in sql
# Since it wont have too much add/drops within tables, 
# confirmation question is not set
# (wont add records in `invoicDate` table currently)

import mysql.connector
import openpyxl
import win32com.client

xl = win32com.client.Dispatch("Excel.Application")
xl.Visible = True
wb_open = True
wb = xl.Workbooks.Open(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\DATABASE_TRIAL\python files\InputInvoice.xlsx")

import os
from dotenv import load_dotenv

load_dotenv()

MYSQL_USER = os.getenv('MYSQL_USER')
MYSQL_PW = os.getenv('MYSQL_PW')

connection = mysql.connector.connect(
    host='localhost',
    port='3306',
    user= MYSQL_USER,
    password= MYSQL_PW,
    database='trial_database'
)



while True:
    confirm = input('請檢查清楚excel表,確認請輸入"Y"\nEXCEL sheet entered? "Y"')
    if confirm == "Y":
        wb.Close(False)
        # below line new added
        xl.Application.Quit()

        excel = openpyxl.load_workbook(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\DATABASE_TRIAL\python files\InputInvoice.xlsx")
        sheet = excel.active


        cursor = connection.cursor()

        # check if invoice in database
        invoiceList = []
        for i in range(2, sheet.max_row + 1):
            if sheet['b' + str(i)].value not in invoiceList:
                invoiceList.append(sheet['b' + str(i)].value)
        print("invoiceList: " + str(invoiceList))
        for inv in invoiceList:
            if inv != None:
                cursor.execute("""
                select * from `refInvoiceNo`
                where `發票` = '{}';
                """.format(inv))
                result = cursor.fetchall()
                if result != []:
                    choose = input("存在相同invoice, 請改一個新invoice 並重新輸入\n或輸入'continue'來追加同一invoice的庫存: ")
                    if choose != 'continue':
                        quit()

        for i in range(2, sheet.max_row + 1):
            ref = sheet['a' + str(i)].value
            if ref == None:
                print(i, sheet.max_row)
                break
            invoice = sheet['b' + str(i)].value
            num = int(sheet['c' + str(i)].value)
            # 現退 = sheet['d' + str(i)].value
            cost = float(sheet['d' + str(i)].value)
            csorone = sheet['e' + str(i)].value
            if invoice == None or csorone == None:
                print("請填寫所有空格")
                import time
                time.sleep(3)
                quit()
                


            cursor.execute("""
            insert into `refInvoiceNo` (`型號`, `發票`, `數量`, `現貨/退貨/盒`, `cost`, `cs/onera`, `入貨日期`) 
            values ('{}', '{}', '{}', '現貨', '{}', '{}', curdate());
                        
            """.format(
                ref, invoice, num, cost, csorone
            ))

        print("成功輸入庫存\nsucceeded inputting stock!!")
            

        excel.save(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\DATABASE_TRIAL\python files\InputInvoice.xlsx")
        
        cursor.close()
        connection.commit()
        connection.close()

        import RefreshStock
        RefreshStock
        quit()
    else:
        print('wrong input!!')


    
