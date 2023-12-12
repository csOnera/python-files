import mysql.connector
import openpyxl
import win32com.client
import os


import os
from dotenv import load_dotenv

load_dotenv()

MYSQL_USER = os.getenv('MYSQL_USER')
MYSQL_PW = os.getenv('MYSQL_PW')

connection = mysql.connector.connect(
    host='localhost',
    port='3306',
    user= MYSQL_USER,
    password= MYSQL_PW,
    database='trial_database'
)

cursor = connection.cursor()


xl = win32com.client.Dispatch("Excel.Application")
xl.Visible = True
wb_open = True
wb = xl.Workbooks.Open(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\DATABASE_TRIAL\python files\input退貨.xlsx")

confirm = input(' EXCEL sheet entered? "Y"')

if confirm == 'Y':
    wb.Close(True)
    excel = openpyxl.load_workbook(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\DATABASE_TRIAL\python files\input退貨.xlsx")
    sheet = excel.active


    numOfRecord = 0
    # for each record
    for j in range(2, sheet.max_row + 1):
        ref = sheet['a' + str(j)].value
        
        if ref == None:
            break
        numOfRecord += 1
        num = int(sheet['b' + str(j)].value)
        inputInvoice = sheet['c' + str(j)].value
        outputInvoice = sheet['d' + str(j)].value

        # select -refId-(only exportId is enough) and exportId
        # cursor.execute("""
        #     select * from refInvoiceNo
        #     where 型號 = '{}';
        # """.format(ref))

        # refRecords = cursor.fetchall()
        # os.system('cls')
        # print(inputInvoice)
        # for i in refRecords:
        #     print(i)

        # refId = input('input the refId (or "null" if not found)...  ')

        cursor.execute("""
            select `id`, `發票`, `數量`, `去處`, `ref_id` from exportRecord
            where 型號 = '{}' and `數量` <> 0;
        """.format(ref))

        exportRecords = cursor.fetchall()
        os.system('cls')
        print(ref, num, outputInvoice)
        idEnsureList = ["null"]

        print("id", "發票", "數量", "去處", "ref_id")
        for i in exportRecords:
            idEnsureList.append(str(i[0]))
            print(i)

        exportId = 0

        while exportId not in idEnsureList:
            exportId = input('input the exportId (or "null" if not found)... \nMake sure type the id shown')

        os.system('cls')
        if exportId != "null":
            
            # add constrain to prohibit getting num > 出貨數量 (but wt if need two sources to 退!!)
            # 整一個loop 比 this case
            cursor.execute("""select * from exportRecord where id = '{}';""".format(exportId))
            result = cursor.fetchall()[0]
            if result[4] < num:
                print('選取的出貨紀錄數量不足, 請核對數量以繼續\n當前退回數量為: {}'.format(num))
                while num > 0:
                    print('選取的出貨紀錄數量不足, 開始迴圈\n現在剩餘數量: {}'.format(num))

                    cursor.execute("""
                        select `id`, `發票`, `數量`, `去處`, `ref_id` from exportRecord
                        where 型號 = '{}' and `數量` <> 0;
                    """.format(ref))

                    exportRecords = cursor.fetchall()
                    print(ref, num, outputInvoice)

                    print("id", "發票", "數量", "去處", "ref_id")
                    for i in exportRecords:
                        print(i)

                    exportId = 0

                    while exportId not in idEnsureList:
                        exportId = input('input the exportId (or "null" if not found)... \nMake sure type the id shown')

                    idEnsureList.remove(exportId)
                    cursor.execute("""select * from exportRecord where id = '{}';""".format(exportId))
                    result = cursor.fetchall()[0]

                    # here minus num export record (cases: recordNum > num, recordNum <= num)
                    # set minusNum and real minus num tends to 0
                    if result[4] >= num:
                        minusNum = num
                        num = 0
                    else:
                        minusNum = result[4]
                        num -= result[4]
                        numOfRecord += 1
                    refId = result[7]

                    # below copy from below low---------------------------------------

                    
                    cursor.execute("""select `發票`, `cs/onera`, `cost` from refInvoiceNo where id = '{}';""".format(refId))
                    result = cursor.fetchall()
                    invoiceFromRefId = result[0][0]
                    csOrOnera = result[0][1]
                    costFromRefId = result[0][2]
                    # edit stock (if already hv new_refId then dun append)
                    # search if there is 退貨, same ref and invoice
                    cursor.execute("""
                        select * from refInvoiceNo
                        where 型號 = '{}'and 發票 = '{}' and `現貨/退貨/盒` = '退貨';
                    """.format(ref, invoiceFromRefId))

                    result = cursor.fetchall()
                    needNewInvoiceRecord = True

                    if result == []:
                        cursor.execute("""
                            insert into refInvoiceNo (`型號`, `數量`, `發票`, `現貨/退貨/盒`, `cost`, `cs/onera`, `入貨日期`)
                            values ('{}', '{}', '{}', '退貨', '{}', '{}', curdate())
                        """.format(ref, minusNum, invoiceFromRefId, costFromRefId, csOrOnera))
                        needNewInvoiceRecord = True
                    # else: already has 
                    else:
                        originalBackId = result[0][0]
                        cursor.execute("""
                            update refInvoiceNo
                            set 數量 = {}
                            where id = '{}';
                        """.format(result[0][3] + minusNum, originalBackId))
                        needNewInvoiceRecord = False
                        

                    # edit exportRecord (+x)
                    # take record info first
                    cursor.execute("""
                        select * from exportRecord
                        where id = '{}';
                    """.format(exportId))
                    result = cursor.fetchall()
                    
                    # update
                    cursor.execute("""
                        update exportRecord
                        set 數量 = {}
                        where id = '{}';
                    """.format(result[0][4] - minusNum, exportId))


                    # edit 退貨紀錄
                    cursor.execute("""
                        select max(`id`) from refInvoiceNo;
                    """)
                    maxid = cursor.fetchall()[0][0]
                    print(maxid)

                    if needNewInvoiceRecord:
                        cursor.execute("""
                            insert into `退貨紀錄` (input_date, 入貨單號, 出貨單號, 型號, backNum, refId, exportId, new_refId)
                            values (curdate(), '{}', '{}', '{}', '{}', '{}', '{}', '{}');
                        """.format(invoiceFromRefId, result[0][5], ref, minusNum, int(refId), int(exportId), maxid))
                    else:
                        cursor.execute("""
                            insert into `退貨紀錄` (input_date, 入貨單號, 出貨單號, 型號, backNum, refId, exportId, new_refId)
                            values (curdate(), '{}', '{}', '{}', '{}', '{}', '{}', '{}');
                        """.format(invoiceFromRefId, result[0][5], ref, minusNum, int(refId), int(exportId), originalBackId))
            
            else:
                refId = result[7]
                cursor.execute("""select `發票`, `cs/onera`, `cost` from refInvoiceNo where id = '{}';""".format(refId))
                result = cursor.fetchall()
                invoiceFromRefId = result[0][0]
                csOrOnera = result[0][1]
                costFromRefId = result[0][2]
                # edit stock (if already hv new_refId then dun append)
                # search if there is 退貨, same ref and invoice
                cursor.execute("""
                    select * from refInvoiceNo
                    where 型號 = '{}'and 發票 = '{}' and `現貨/退貨/盒` = '退貨';
                """.format(ref, invoiceFromRefId))

                result = cursor.fetchall()
                needNewInvoiceRecord = True

                if result == []:
                    cursor.execute("""
                        insert into refInvoiceNo (`型號`, `數量`, `發票`, `現貨/退貨/盒`, `cost`, `cs/onera`, `入貨日期`)
                        values ('{}', '{}', '{}', '退貨', '{}', '{}', curdate())
                    """.format(ref, num, invoiceFromRefId, costFromRefId, csOrOnera))
                    needNewInvoiceRecord = True
                # else: already has 
                else:
                    originalBackId = result[0][0]
                    cursor.execute("""
                        update refInvoiceNo
                        set 數量 = {}
                        where id = '{}';
                    """.format(result[0][3] + num, originalBackId))
                    needNewInvoiceRecord = False
                    

                # edit exportRecord (+x)
                # take record info first
                cursor.execute("""
                    select * from exportRecord
                    where id = '{}';
                """.format(exportId))
                result = cursor.fetchall()
                
                # update
                cursor.execute("""
                    update exportRecord
                    set 數量 = {}
                    where id = '{}';
                """.format(result[0][4] - num, exportId))


                # edit 退貨紀錄 (stopped here)
                cursor.execute("""
                    select max(`id`) from refInvoiceNo;
                """)
                maxid = cursor.fetchall()[0][0]
                print(maxid)

                if needNewInvoiceRecord:
                    cursor.execute("""
                        insert into `退貨紀錄` (input_date, 入貨單號, 出貨單號, 型號, backNum, refId, exportId, new_refId)
                        values (curdate(), '{}', '{}', '{}', '{}', '{}', '{}', '{}');
                    """.format(invoiceFromRefId, result[0][5], ref, num, int(refId), int(exportId), maxid))
                else:
                    cursor.execute("""
                        insert into `退貨紀錄` (input_date, 入貨單號, 出貨單號, 型號, backNum, refId, exportId, new_refId)
                        values (curdate(), '{}', '{}', '{}', '{}', '{}', '{}', '{}');
                    """.format(invoiceFromRefId, result[0][5], ref, num, int(originalBackId), int(exportId), originalBackId))

        # here if exportId == "null"
        else:
            # add record to stock
            cursor.execute("""
                insert into refInvoiceNo (`型號`, `數量`, `發票`, `現貨/退貨/盒`, `cs/onera`, `入貨日期`)
                values ('{}', '{}', '{}', '退貨', 'cs', curdate())
            """.format(ref, num, inputInvoice))

            cursor.execute("""
                select max(`id`) from refInvoiceNo;
            """)
            maxid = cursor.fetchall()[0][0]
            # print(maxid)

            # add record to 退貨表
            cursor.execute("""
                insert into `退貨紀錄` (input_date, 入貨單號, 出貨單號, 型號, backNum, refId, new_refId)
                values (curdate(), '{}', '{}', '{}', '{}', '{}', '{}');
            """.format(inputInvoice, outputInvoice, ref, num, maxid, maxid))
        
    print("database done! ")
    connection.commit()


    eer = openpyxl.load_workbook(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\DATABASE_TRIAL\python files\exportExportRecords.xlsm", read_only=False, keep_vba=True)
    eerws = eer.active
    eerws.delete_rows(2, sheet.max_row)
    # xl.Application.Run('exportExportRecords.xlsm!Module5.ClearRows')

    # export the last 退貨紀錄
    cursor.execute("""
        select `入貨單號`, `型號`, `backNum`, `出貨單號` from `退貨紀錄`
        order by `id` desc limit {};
    """.format(numOfRecord))
    result = cursor.fetchall()

    for row in range(2, len(result) + 2):
        eerws["c" + str(row)].value = result[row-2][1]
        eerws["d" + str(row)].value = result[row-2][0]
        eerws["e" + str(row)].value = result[row-2][2]
        eerws["f" + str(row)].value = result[row-2][3] + "退貨 +"

    eer.save(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\DATABASE_TRIAL\python files\exportExportRecords.xlsm")
    from exportExport import runVBA

    runVBA()

    import RefreshStock
    RefreshStock
    import RefreshRecord
    RefreshRecord


cursor.close()
connection.commit()
connection.close()