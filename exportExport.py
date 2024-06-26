# this program export export records from sql to excel and will renew 
# excel worksheet everytimes running
# 2nd part: entering 'C' to continue auto filling records in CS 出入貨
# newly opened excel don't have macro thus 2nd part






def exportExport(Q, NumItem = 0, skipAskingId = 0):
    import mysql.connector
    import openpyxl

    if Q == "num" and NumItem == 0:
        NumItem = int(input('No. of items (integer)'))
    if Q == "id":
        if skipAskingId == 0:
            id = int(input('id (integer)'))  
        else:
            id = skipAskingId
        NumItem = int(input('No. of items (integer)'))
        if id <= 2850:
            print("No access to previous records")
            while True:
                toquit = input("type 'quit' to quit: ")
                if toquit == "quit":
                    quit()
    try:
        excel = openpyxl.load_workbook(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\DATABASE_TRIAL\python files\exportExportRecords.xlsm", read_only=False, keep_vba=True)
        sheet = excel.active
        # sheet.title = 'sheet2'
        # excel.create_sheet('Sheet')

        # # excel.remove(sheet)
        # import win32com.client
        # xl = win32com.client.Dispatch("Excel.Application")
        # xl.Application.Run('exportExportRecords.xlsm!Module5.ClearRows')
        sheet.delete_rows(2, sheet.max_row)

    except:
        # excel = openpyxl.Workbook()
        print("請先關閉exportExportRecords 的excel檔案\nplease close exportExportRecords before running the app")
        import time
        time.sleep(5)
        quit()

    sheet = excel.active

    import os
    from dotenv import load_dotenv

    load_dotenv()

    MYSQL_USER = os.getenv('MYSQL_USER')
    MYSQL_PW = os.getenv('MYSQL_PW')

    connection = mysql.connector.connect(
        host='localhost',
        port='3306',
        user= MYSQL_USER,
        password= MYSQL_PW,
        database='trial_database'
    )

    cursor = connection.cursor()


    # START HERE   
    cursor.execute("""
        show columns from `exportRecord`;
    """)
    records = cursor.fetchall()
    records.append(["cost"])
    records.append(["現貨/退貨/盒"])
    colNum = len(records) 
    for i in records:
        sheet.cell(row = 1, column = records.index(i) + 1).value = i[0]
    

    if Q == "num":
        cursor.execute("""
            select `exportRecord`.*, `cost`, `現貨/退貨/盒` from `exportRecord`
            join `refInvoiceNo`
            on `ref_id` = `refInvoiceNo`.`id`
            order by `exportRecord`.`id` desc limit {};
        """.format(NumItem))
    elif Q == "id":
        cursor.execute("""
            select `exportRecord`.*, `cost`, `現貨/退貨/盒` from `exportRecord`
            join `refInvoiceNo`
            on `ref_id` = `refInvoiceNo`.`id`
            where `exportRecord`.`id` >= '{}'
            limit {};
        """.format(id, NumItem))

    records = cursor.fetchall()
    # here reverse if num for better copying
    if Q == "num":
        records.reverse()

    for i in range(1,len(records)+1):
        for j in range(colNum):
            sheet.cell(row = i + 1, column = j + 1).value = records[i - 1][j]



    print("成功 succeeded!")

    excel.save(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\DATABASE_TRIAL\python files\exportExportRecords.xlsm")

    cursor.close()
    connection.commit()
    connection.close()

    

    
def runVBA():
    import win32com.client
    from fileOperation import read_file
    xl = win32com.client.Dispatch("Excel.Application")
    xl.Visible = True
    wb = xl.Workbooks.Open(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\DATABASE_TRIAL\python files\exportExportRecords.xlsm")

    confirm = input("請檢查清楚從數據庫中導出的excel檔案的資料,如無誤可輸入'C'以自動填入出貨紀錄至出入貨excel表\nCHECK THE EXCEL SHEET & 'C' to auto plug-in records to CS 出入貨 ")

    if confirm == 'C':
        year = read_file()
        print(year)
        # xl.Workbooks.Open("C:/Users/onera/OneDrive - ONE ERA (HK) LIMITED/oneraShare/CHARMSMART " + year + " 出入貨紀錄.xlsx")
        # xl.Workbooks.Open(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\以前紀錄\CHARMSMART 2022 財政年度 出·入貨表ver2.xlsx")

        xl.Application.Run('exportExportRecords.xlsm!Module1.autoExportRecord')

        wb.Close(True)
        # xl.Application.Quit()



# exportExport("id")

# runVBA()