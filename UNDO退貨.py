# to revert input from 入退貨

import openpyxl
import mysql.connector
import win32com.client
from exportExport import exportExport, runVBA


import os
from dotenv import load_dotenv

load_dotenv()

MYSQL_USER = os.getenv('MYSQL_USER')
MYSQL_PW = os.getenv('MYSQL_PW')

connection = mysql.connector.connect(
    host='localhost',
    port='3306',
    user= MYSQL_USER,
    password= MYSQL_PW,
    database='trial_database'
)
cursor = connection.cursor()

# two parts
# first to add back the number in `refInvoiceNo`
# get the invoice and ref from `exportRecords`


ref = input('請輸入要查找型號或輸入"quit"退出查找')

if ref == "quit":
    quit()
else:
    print("id, input_date, 入貨單號, 出貨單號, 型號, backNum, refId, exportId, new_refId")
    idList = []
    cursor.execute("""
        select * from `退貨紀錄`
        where `型號` = '{}';
    """.format(ref))
    list = cursor.fetchall()
    if list == []:
        print('庫存沒有該型號的退貨記錄\n')
    for i in list:
        print(i)
        idList.append(i[0])
undoId = 0

while undoId not in idList:
    undoId = int(input("id (integer)"))

amoutOfUndo = int(input("例子: 如要undo id為301,302,303,304的紀錄則先輸入301, 現在輸入退貨項目數量輸入 4(四個紀錄)\n請輸入要UNDO的項目數量\nnumber of items to UNDO"))

cursor.execute("""
    select * from `退貨紀錄`
    where `id` >= '{}' limit {};
""".format(undoId, amoutOfUndo))

result = cursor.fetchall()

# here print the export data in 退貨
backExcelPath = r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\DATABASE_TRIAL\python files\input退貨.xlsx"
backExcel = openpyxl.load_workbook(backExcelPath)
backws = backExcel.active
backws.delete_rows(2, backws.max_row)

for row in range(2, 2 + len(result)):
    backws["a" + str(row)].value = result[row-2][4]
    backws["b" + str(row)].value = result[row-2][5]
    backws["c" + str(row)].value = result[row-2][2]
    backws["d" + str(row)].value = result[row-2][3]
    backws["e" + str(row)].value = result[row-2][0]
    backws["f" + str(row)].value = result[row-2][7]
    backws["g" + str(row)].value = result[row-2][8]

backExcel.save(backExcelPath)
print('loading data ...')


xl = win32com.client.Dispatch("Excel.Application")
xl.Visible = True

wb_open = True
wb = xl.Workbooks.Open(backExcelPath)

checked = input('請檢查清楚excel檔的資料, 如無問題可輸入"checked"來還原出貨紀錄\nCHECK items to be reverted in the input退貨.xlsm "checked"')
# while True:
if checked == "checked" :
    wb.Close(False)
    excel = openpyxl.load_workbook(backExcelPath)
    sheet = excel.active
    # loop for >1 records to be reverted in .xlsm
    for i in range(2, sheet.max_row + 1):
        exportId = sheet["f" + str(i)].value
        numAddBack = sheet["b" + str(i)].value
        newRefId = sheet["g" + str(i)].value
        backId = sheet["e" + str(i)].value


        # take the no. in stock first (big bug happened: more than one record of same ref same invoice)
        # should take ref_id
        cursor.execute("""
            select `數量` from `refInvoiceNo`
            where `id` = '{}';
        """.format(newRefId))
        NumInStock = cursor.fetchall()[0][0]
        cursor.execute("""
            update `refInvoiceNo`
            set `數量` = '{}'
            where `id` = '{}';
        """.format(NumInStock - numAddBack, newRefId))

        # add back exportRecord number
        cursor.execute("""
            select `數量` from `exportRecord`
            where `id` = '{}';
        """.format(exportId))

        try:
            NumInExport = cursor.fetchall()[0][0]
        
            cursor.execute("""
                update `exportRecord`
                set `數量` = '{}'
                where `id` = '{}';
            """.format(NumInExport + numAddBack, exportId))
        except:
            print("冇相關出貨紀錄\nno export record found")
            pass
        
        # last to delete the records in `退貨紀錄`
        cursor.execute("""
            delete from `退貨紀錄`
            where `id` = '{}'
        """.format(backId))


print("done!")
xl.Application.Quit()

import RefreshStock
RefreshStock
import RefreshRecord
RefreshRecord

cursor.close()
connection.commit()
connection.close()