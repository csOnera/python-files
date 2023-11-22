import openpyxl

tbaddedAddress = r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\得物對賬\23 3.01-5.28 跨境寄售订单列表1679627165290.xlsx"

refSheet = openpyxl.load_workbook(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\得物對賬\總比對對賬單&交易成功.xlsx", data_only=True)
ToBeAdded = openpyxl.load_workbook(tbaddedAddress)
# dewu = openpyxl.load_workbook(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\出貨OR退貨紀錄\得物出入貨(含約倉).xlsx")

tobeaddedsheet = ToBeAdded.active
# dewuSheet = dewu.active

def search(ap, ref):
    if ap in refSheet.sheetnames:
        # locate the worksheet
        for i in range(len(refSheet.sheetnames)):
            if ap == refSheet.sheetnames[i]:
                refSheet.active = i
                refBook = refSheet.active
                break
        # find the price of that ref by column m and n
        for i in range(2,100):
            if refBook["n" + str(i)].value == ref:
                return refBook["m" + str(i)].value
        #  0 represents ref not found in that ap
        return 0
    else:
        #  1 represents ap not found
        return 1
    # else:
    #     for i in range(2374, dewuSheet.max_row):
    #         if ap == dewuSheet["b" + str(i)].value:
    #             for j in range(i, dewuSheet.max_row):
    #                 if ref == dewuSheet["f" + str(i)].value
        

for i in range(2, tobeaddedsheet.max_row+1):
    # column shd be AP/CAP
    if tobeaddedsheet["q" + str(i)].value != None:
        # AP/CAP
        refNum = tobeaddedsheet["q" + str(i)].value
        # Txxx.x... (possible error: last chinese character)
        ref = tobeaddedsheet["e" + str(i)].value
        price = search(refNum,ref)
        # column where to insert
        tobeaddedsheet["m" + str(i)].value = price


ToBeAdded.save(tbaddedAddress)

print("done!")
