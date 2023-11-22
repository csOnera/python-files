import pathlib

folder = pathlib.Path(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\報關單2022")
flist = list(folder.iterdir())

# import openpyxl
# print(flist[0])
# samplewb = openpyxl.load_workbook(flist[0])


# need to turn .xls to .xlsx files
import win32com.client as win32

# print(len(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\xlsx\\" + flist[1].name + 'x'))
import pandas as pd
import pyexcel
# print(flist[3])

# pyexcel.save_book_as(file_name="C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\報關單2022\郑州单据BIP1194705709-1194704724.xls",dest_file_name="C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\xlsx\\郑州单据BIP1194705709-1194704724.xlsx")

def checkInCopied(x):
    xlsxFolder = pathlib.Path(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\xlsx")
    xlsxList = list(xlsxFolder.iterdir())
    for i in xlsxList:
        if x == i.name:
            return True
    return False


# not sure if succeeded: only one sheet copied, not certain spacing;
# bug to be fixed: file name in flist in .xlsx instead of .xls
# xls = str(flist[0]).replace('xlsx','xls')
# print(xls, str(flist[0]))
# for i in flist:
#     print(i.name)
#     if checkInCopied(i.name) == False and str(i)[-4:] == 'xlsx':
#         df = pd.read_excel(str(i))
#         df.to_excel(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\xlsx\\" + i.name)

# then to add cost
xlsxFolder = pathlib.Path(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\xlsx")
xlsxList = list(xlsxFolder.iterdir())
import openpyxl

# cs2021 = openpyxl.load_workbook(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\以前紀錄\charm2021出·入貨 latest version-DESKTOP-833R29B.xlsx")
cs2022 = openpyxl.load_workbook(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\以前紀錄\CHARMSMART 2022 財政年度 出·入貨表ver2.xlsx")
cs2023 = openpyxl.load_workbook(r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\oneraShare\CHARMSMART 2023 出入貨紀錄.xlsx")
# cs2021s = cs2021.active
cs2022s = cs2022.active
cs2023s = cs2023.active


for i in range(len(xlsxList)):
    print(i, xlsxList[i].name)

# first get the bip without 'bip'
# print(xlsxList[0].name)
for k in range(15, len(xlsxList)):
    samplewb = openpyxl.load_workbook(xlsxList[k])
    samplews = samplewb.active
    bipDict = {}
    for i in range(1, samplews.max_row):
        for bipcol in range(5, samplews.max_column):
            if type(samplews.cell(row=i , column=bipcol).value) == str:
                if "BIP" in samplews.cell(row=i ,column=bipcol).value:
                    maxCol = chr(bipcol + 96)
                    break
            

    # maxCol = chr(samplews.max_column + 96)
    for i in range(1, samplews.max_row):
        if type(samplews[maxCol + str(i)].value) == str:
            if samplews[maxCol + str(i)].value.replace('BIP', "") not in bipDict and samplews[maxCol + str(i)].value[0:3] == "BIP":
                bipDict[samplews[maxCol + str(i)].value.replace('BIP', "")] = {}
    
    print(xlsxList[k].name)


    # then to extract all the bip in cs出入


    # cs2022 first
    # keyList = list(bipDict.keys)
    for bip in bipDict.keys():
        print(bip)
        for y in range(1,cs2023s.max_row):
            for x in range(10, cs2023s.max_column):
                if type(cs2023s[chr(x + 96) + str(y)].value) == str:
                    if str(bip) in cs2023s[chr(x + 96) + str(y)].value:
                        print(cs2023s[chr(x + 96) + str(y)].value)
                        bipDict[bip][cs2023s['d' + str(y)].value] = cs2023s['h' + str(y)].value
    print('cs 2022 records imported dict')

    import re

    # insert back the cost
    for i in range(1, samplews.max_row):
        if samplews[maxCol + str(i)].value not in bipDict and samplews[maxCol + str(i)].value != None and samplews[maxCol + str(i)].value[0:3] == "BIP":
            noBip = samplews[maxCol + str(i)].value.replace('BIP', "")
            ref = re.search('[a-zA-Z]+[\d.]+',samplews['e' + str(i)].value).group()
            print(ref)
            if samplews['n' + str(i)].value == None:
                samplews['n' + str(i)].value = bipDict[noBip].get(ref)
            else:
                print('column n not none')



    samplewb.save(xlsxList[k])