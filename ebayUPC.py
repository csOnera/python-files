# TURNS OUT CERTINA AND SWATCH NOT FOUND IN EBAY LOL 

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
# from webdriver_manager.chrome import ChromeDriverManager

from selenium.webdriver.common.by import By

from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
# import selenium.webdriver.support.ui as ui

# from webdriver_manager.chrome import ChromeDriverManager
import time


import openpyxl

import re

# ADD ID, MONEY
service = Service()
options = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=options)

serialPath = r"C:\Users\onera\OneDrive - ONE ERA (HK) LIMITED\雜\operate Serial No. by Scanning.xlsm"

wb = openpyxl.load_workbook(serialPath, read_only=False, keep_vba=True)
ws = wb.active

count = 0
for row in range(1, ws.max_row):
    if ws["a" + str(row)].value != None:
        count += 1
        driver.get(f'https://www.ebay.ca/sch/i.html?_from=R40&_trksid=p4432023.m570.l1313&_nkw={ws["a" + str(row)].value}+&_sacat=0')

        try:
            # element = WebDriverWait(driver, 5).until(
            # EC.presence_of_element_located(By.CLASS_NAME, 's-item__link'))
            time.sleep(2)
            # print('not wait error')
            title = driver.find_element(By.XPATH, '/html/body/div[5]/div[4]/div[2]/div[1]/div[2]/ul/li[2]/div/div[2]/a/div/span').text
            print(title)
            if type(title) == str:
                ws["b" + str(row)].value = re.search('[A-Z][\d\.]+',title).group()
            else:
                 print('type not string')
        except:
            print('error')
        finally:
            print(count)
    # if count == 6:
    #     break
    

wb.save(serialPath)

while True:
        x = input('quit to quit')
        if x.lower() == 'quit':
            driver.quit()
            exit()